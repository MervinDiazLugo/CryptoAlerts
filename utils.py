import smtplib
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.utils import COMMASPACE, formatdate
from openpyxl import load_workbook
import requests
import json
from datetime import datetime, timedelta
from os.path import basename

PLAZO = 10
PORCENTAJE = 1.8
DESDE = (datetime.utcnow() - timedelta(days = PLAZO)).timestamp()
HASTA  = datetime.utcnow().timestamp()
EXCEL_FILE ='./template/Crypto.xlsx'
EXCEL_FILE_OUTPUT ='./output/Crypto_Copy.xlsx'
COIN_LIST = []
retrieve_coin={}
ESPERADO = ""

class Utils:

    def workbook(self, excel_file=EXCEL_FILE):
        global wb
        wb = load_workbook(excel_file)
        return wb

    def load_sheet(self, excel_file=EXCEL_FILE, hoja="Hoja1"):
        excel = Utils.workbook(excel_file)
        ws = excel[hoja]
        return ws

    def retrieve_coin_trend(self):
        coin_trend = json.loads(requests.get("https://api.coingecko.com/api/v3/search/trending").text)
        coin_trend = coin_trend['coins']
        _coin_trend = []
        for _trend in coin_trend:
            items = _trend['item']
            _coin_trend.append(items['id'])
        print(_coin_trend)
        return _coin_trend

    def retrieve_coin_estimated_volume(self, id):
        total_volumes = json.loads(requests.get(
            f"https://api.coingecko.com/api/v3/coins/{id}/market_chart?vs_currency=usd&days={PLAZO}&interval=daily").text)

        avg = 0
        most = 0
        volume = total_volumes['total_volumes']
        count = 0

        for vol in volume:
            if vol[1] > most:
                most = vol[1]

        for vol in volume:
            avg = avg + vol[1]
            count = count + 1

        avg = avg / count
        avg = (avg+most)/2

        return avg

    def retrieve_coin_data(self, excel_file=EXCEL_FILE, hoja="Hoja1",  estimaciones=[]):
        sheet = self.load_sheet(excel_file, hoja)
        row = 2
        count = 0
        for coin in estimaciones:
            try:
                coin_data = json.loads(requests.get(
                    "https://api.coingecko.com/api/v3/coins/markets?vs_currency=usd&ids={}&order=market_cap_desc&per_page=100&page=1&sparkline=false".format(
                        coin)).text)
                coin_data = coin_data[0]
                retrieve_coin["id"] =coin_data['id']
                retrieve_coin["symbol"] = coin_data['symbol']
                retrieve_coin["current_price"] = coin_data['current_price']
                retrieve_coin["market_cap"] = coin_data['market_cap']
                retrieve_coin["circulating_supply"] = coin_data['circulating_supply']
                retrieve_coin["total_volume"] = coin_data['total_volume']
                retrieve_coin["real_market_price"] = coin_data['market_cap'] / coin_data['circulating_supply']
                retrieve_coin["diferencial"] = retrieve_coin["real_market_price"] - coin_data['current_price']
                retrieve_coin["volumen_avg"] = Utils.retrieve_coin_estimated_volume(self, coin_data['id'])

                if coin_data['total_volume'] >= (retrieve_coin["volumen_avg"] * 1.2):
                    retrieve_coin["expectativa"] = "ALTA"
                else:
                    retrieve_coin['expectativa'] = "BAJA"

                barrera = float(retrieve_coin["volumen_avg"]*(PORCENTAJE))
                if retrieve_coin["total_volume"] >= barrera:
                    retrieve_coin['Alerta'] = "ALERTA"
                else:
                    retrieve_coin['Alerta'] = ""
                print(retrieve_coin)

                for col, value in enumerate(retrieve_coin.values(), start=1):
                    sheet.cell(row=row, column=col).value = value
                    sheet.cell(row=row, column=col).number_format = '#,##0.0000000000000'
                wb.save(EXCEL_FILE_OUTPUT)
                row = row+1
                count = count + 1
                retrieve_coin.clear()
            except (IndexError, KeyError, ValueError, ZeroDivisionError, TypeError) as error:
                print(error, coin)
                continue
        Utils.mailing(self)


    def retrieve_coin_list(self):
        _coin_list = json.loads(requests.get("https://api.coingecko.com/api/v3/coins/list?include_platform=false").text)
        coin_list = []
        for coins in _coin_list:
            coin_list.append(coins['id'])
        print(coin_list)
        return coin_list

    ##########################################################################
    #########################   -=_EMAIL_=-   #########################
    ##########################################################################
    def send_mail(subject, text, files=None):
        strTimeSubject = str(datetime.today().strftime("%Y-%m-%dT%H:%M:%S"))
        msgAsunto = "Crypto Alerts: " + strTimeSubject
        msg = MIMEMultipart()
        msg['From'] = "mervindiazlugo@pepisandbox.com"
        msgTO = ["mervindiazlugo@gmail.com"]
        msg['To'] = COMMASPACE.join(msgTO)
        msg['Date'] = formatdate(localtime=True)
        msg['Subject'] = msgAsunto

        msg.attach(MIMEText(text))

        for f in files or []:
            with open(f, "rb") as fil:
                part = MIMEApplication(
                    fil.read(),
                    Name=basename(f)
                )
            # After office close
            part['Content-Disposition'] = 'attachment; filename="%s"' % basename(f)
            msg.attach(part)

        smtpHost = "smtp.pepipost.com"
        smtpPort = "25"
        smtpUser = "mervindiazlugo"
        smtpPass = "mervindiazlugo_1243b9480a64d0fd9a2ec1e3c8e58b74"

        smtp = smtplib.SMTP(smtpHost, smtpPort)
        smtp.set_debuglevel(False)
        smtp.login(smtpUser, smtpPass)
        smtp.default_port
        smtp.sendmail(msg['From'], msg['To'], msg.as_string())
        smtp.close()

    def mailing(self, Adjuntos=EXCEL_FILE_OUTPUT):
        msgBody = ('''
                            Esta notificacion fue generada de forma automatica, al concluir el proceso pruebas automatizado

                            --> (#) Adjuntos los indicadores correspondientes. (#) <--


                            ''')
        #
        msgBody = msgBody.format()
        msgAttachments = [Adjuntos]
        Utils.send_mail(self, msgBody, msgAttachments)
